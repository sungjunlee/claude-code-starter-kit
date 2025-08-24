#!/usr/bin/env -S uv run
# /// script
# dependencies = [
#     "pypandoc",
#     "markdown2",
#     "jinja2",
#     "openpyxl",
#     "beautifulsoup4",
#     "click",
# ]
# requires-python = ">=3.10"
# ///

"""
Document Converter for Research Assistant
Converts markdown research reports to various formats (PDF, DOCX, XLSX)
"""

import click
import os
import sys
import webbrowser
import tempfile
import re
from pathlib import Path
from datetime import datetime
from typing import Optional, List, Dict

import markdown2
from jinja2 import Template, FileSystemLoader, Environment
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Get script directory for templates and styles
SCRIPT_DIR = Path(__file__).parent.resolve()
TEMPLATE_DIR = SCRIPT_DIR.parent / "templates"
STYLE_DIR = SCRIPT_DIR.parent / "styles"

# Default HTML template if no template file exists
DEFAULT_TEMPLATE = """
<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{{ title }}</title>
    <style>
        /* Basic styling for professional documents */
        @import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;700&display=swap');
        
        body {
            font-family: 'Noto Sans KR', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
            line-height: 1.6;
            color: #333;
            max-width: 900px;
            margin: 0 auto;
            padding: 20px;
        }
        
        h1 { 
            color: #2c3e50;
            border-bottom: 3px solid #3498db;
            padding-bottom: 10px;
        }
        
        h2 {
            color: #34495e;
            margin-top: 30px;
            border-bottom: 1px solid #ecf0f1;
            padding-bottom: 5px;
        }
        
        h3 { color: #7f8c8d; }
        
        table {
            border-collapse: collapse;
            width: 100%;
            margin: 20px 0;
        }
        
        th, td {
            border: 1px solid #ddd;
            padding: 12px;
            text-align: left;
        }
        
        th {
            background-color: #3498db;
            color: white;
            font-weight: bold;
        }
        
        tr:nth-child(even) {
            background-color: #f9f9f9;
        }
        
        blockquote {
            border-left: 4px solid #3498db;
            margin: 20px 0;
            padding-left: 20px;
            color: #555;
            font-style: italic;
        }
        
        code {
            background-color: #f4f4f4;
            padding: 2px 6px;
            border-radius: 3px;
            font-family: 'Courier New', monospace;
        }
        
        pre {
            background-color: #f4f4f4;
            padding: 15px;
            border-radius: 5px;
            overflow-x: auto;
        }
        
        a {
            color: #3498db;
            text-decoration: none;
        }
        
        a:hover {
            text-decoration: underline;
        }
        
        .metadata {
            background-color: #ecf0f1;
            padding: 15px;
            border-radius: 5px;
            margin-bottom: 30px;
        }
        
        .metadata p {
            margin: 5px 0;
        }
        
        @media print {
            body {
                font-size: 11pt;
                line-height: 1.5;
            }
            
            h1 { page-break-after: avoid; }
            h2, h3 { page-break-after: avoid; }
            
            table { page-break-inside: avoid; }
            
            .metadata {
                background-color: #f5f5f5 !important;
                -webkit-print-color-adjust: exact;
                color-adjust: exact;
            }
        }
    </style>
    {% if custom_css %}
    <style>{{ custom_css }}</style>
    {% endif %}
</head>
<body>
    {% if metadata %}
    <div class="metadata">
        {% if metadata.title %}<p><strong>제목:</strong> {{ metadata.title }}</p>{% endif %}
        {% if metadata.date %}<p><strong>작성일:</strong> {{ metadata.date }}</p>{% endif %}
        {% if metadata.author %}<p><strong>작성자:</strong> {{ metadata.author }}</p>{% endif %}
        {% if metadata.agents %}<p><strong>사용 에이전트:</strong> {{ metadata.agents }}</p>{% endif %}
        {% if metadata.sources_count %}<p><strong>참조 출처:</strong> {{ metadata.sources_count }}개</p>{% endif %}
    </div>
    {% endif %}
    
    {{ content }}
    
    <footer style="margin-top: 50px; padding-top: 20px; border-top: 1px solid #ddd; color: #666; font-size: 0.9em;">
        <p>생성일: {{ generated_date }}</p>
    </footer>
</body>
</html>
"""


def validate_sources(content: str) -> dict:
    """Validate and analyze source links in the document"""
    import re
    from urllib.parse import urlparse
    
    # Find all markdown links
    link_pattern = r'\[([^\]]+)\]\(([^)]+)\)'
    links = re.findall(link_pattern, content)
    
    source_analysis = {
        'total_links': len(links),
        'valid_domains': 0,
        'government_sources': 0,
        'academic_sources': 0,
        'legal_sources': 0,
        'broken_links': [],
        'quality_score': 0
    }
    
    government_domains = ['go.kr', 'gov', '.gov.', 'moleg.go.kr', 'law.go.kr']
    academic_domains = ['edu', '.ac.kr', 'scholar.google', 'jstor', 'doi.org']
    legal_domains = ['scourt.go.kr', 'court.go.kr', 'lawnet.co.kr', 'glaw.scourt.go.kr']
    
    for text, url in links:
        try:
            parsed = urlparse(url)
            if parsed.scheme and parsed.netloc:
                source_analysis['valid_domains'] += 1
                
                domain = parsed.netloc.lower()
                if any(gov in domain for gov in government_domains):
                    source_analysis['government_sources'] += 1
                elif any(edu in domain for edu in academic_domains):
                    source_analysis['academic_sources'] += 1
                elif any(legal in domain for legal in legal_domains):
                    source_analysis['legal_sources'] += 1
        except:
            source_analysis['broken_links'].append(url)
    
    # Calculate quality score
    if source_analysis['total_links'] > 0:
        quality_factors = [
            source_analysis['government_sources'] / source_analysis['total_links'] * 0.4,
            source_analysis['academic_sources'] / source_analysis['total_links'] * 0.3,
            source_analysis['legal_sources'] / source_analysis['total_links'] * 0.3,
            (1 - len(source_analysis['broken_links']) / source_analysis['total_links']) * 0.2
        ]
        source_analysis['quality_score'] = min(100, sum(quality_factors) * 100)
    
    return source_analysis


def enhance_document_quality(content: str, metadata: dict) -> tuple[str, dict]:
    """Enhance document quality with automatic improvements"""
    enhanced_content = content
    quality_report = {}
    
    # 1. Source validation
    source_analysis = validate_sources(content)
    quality_report['sources'] = source_analysis
    
    # 2. Add missing metadata
    if not metadata.get('sources_count'):
        metadata['sources_count'] = source_analysis['total_links']
    
    if not metadata.get('quality_score'):
        metadata['quality_score'] = f"{source_analysis['quality_score']:.1f}%"
    
    # 3. Korean legal citation format improvement
    import re
    
    # Improve Korean court case citations
    case_pattern = r'(\d{4}[가-힣]+\d+)'
    cases = re.findall(case_pattern, enhanced_content)
    if cases:
        quality_report['legal_citations'] = len(cases)
    
    # Add legal document disclaimer if not present
    if not re.search(r'법적\s*조언|법률\s*자문', enhanced_content):
        disclaimer = "\n\n---\n**면책조항**: 본 문서는 연구 및 참고 목적으로 작성되었으며, 구체적인 법적 조언을 구성하지 않습니다. 실무 적용 시 전문가의 자문을 받으시기 바랍니다."
        enhanced_content += disclaimer
    
    return enhanced_content, quality_report


def extract_metadata(markdown_text: str) -> tuple[dict, str]:
    """Extract YAML frontmatter from markdown and enhance with quality data"""
    metadata = {}
    content = markdown_text
    
    # Check for YAML frontmatter
    if markdown_text.startswith('---'):
        try:
            end_index = markdown_text.index('---', 3)
            yaml_text = markdown_text[3:end_index].strip()
            content = markdown_text[end_index + 3:].strip()
            
            # Simple YAML parsing (for basic key-value pairs)
            for line in yaml_text.split('\n'):
                if ':' in line:
                    key, value = line.split(':', 1)
                    metadata[key.strip()] = value.strip()
        except ValueError:
            pass
    
    # Enhance document quality
    enhanced_content, quality_report = enhance_document_quality(content, metadata)
    
    # Add quality information to metadata
    if quality_report:
        metadata['quality_report'] = quality_report
    
    return metadata, enhanced_content


def generate_toc(html_content: str) -> str:
    """Generate a simple table of contents from headers"""
    import re
    
    # Find all headers
    header_pattern = r'<h([1-6])(?:[^>]*id="([^"]*)"[^>]*|[^>]*)>([^<]+)</h[1-6]>'
    headers = re.findall(header_pattern, html_content)
    
    if not headers:
        return ''
    
    toc_html = '<div class="toc"><h2>목차</h2><ul>'
    
    for level, header_id, title in headers:
        level = int(level)
        if level <= 3:  # Only show h1, h2, h3 in TOC
            indent = '  ' * (level - 1)
            # Create anchor if no ID exists
            anchor = header_id if header_id else title.lower().replace(' ', '-')
            toc_html += f'\n{indent}<li><a href="#{anchor}">{title}</a></li>'
    
    toc_html += '\n</ul></div>'
    return toc_html


def markdown_to_html(markdown_file: Path, template_name: Optional[str] = None) -> str:
    """Convert markdown to HTML with template"""
    
    # Read markdown file
    with open(markdown_file, 'r', encoding='utf-8') as f:
        markdown_text = f.read()
    
    # Extract metadata and content
    metadata, content = extract_metadata(markdown_text)
    
    # Process [TOC] before converting to HTML
    toc_placeholder = ""
    if '[TOC]' in content:
        toc_placeholder = "<!--TOC_PLACEHOLDER-->"
        content = content.replace('[TOC]', toc_placeholder)
    
    # Convert markdown to HTML with enhanced features
    html_content = markdown2.markdown(
        content,
        extras=[
            'tables', 
            'fenced-code-blocks', 
            'footnotes', 
            'strike', 
            'task_list',
            'cuddled-lists',    # Better list formatting
            'header-ids',       # Automatic header IDs for navigation
            'wiki-tables'       # Enhanced table support
        ]
    )
    
    # Process Mermaid diagrams - convert ```mermaid blocks to proper divs
    import re
    # markdown2 generates class="language-mermaid" for ```mermaid blocks
    mermaid_pattern = r'<pre><code class="language-mermaid">(.*?)</code></pre>'
    html_content = re.sub(mermaid_pattern, r'<div class="mermaid">\1</div>', html_content, flags=re.DOTALL)
    
    # Generate TOC if placeholder exists
    if toc_placeholder and "<!--TOC_PLACEHOLDER-->" in html_content:
        toc_html = generate_toc(html_content)
        html_content = html_content.replace("<!--TOC_PLACEHOLDER-->", toc_html)
    
    # Load template (default to legal for Korean legal research)
    if not template_name:
        template_name = "legal"  # Default to legal template
        
    if template_name and TEMPLATE_DIR.exists():
        env = Environment(loader=FileSystemLoader(TEMPLATE_DIR))
        try:
            template = env.get_template(f"{template_name}.html")
        except:
            # Fallback to legal if specified template doesn't exist
            try:
                template = env.get_template("legal.html")
            except:
                template = Template(DEFAULT_TEMPLATE)
    else:
        template = Template(DEFAULT_TEMPLATE)
    
    # Load custom CSS if available
    custom_css = ""
    css_file = STYLE_DIR / "professional.css"
    if css_file.exists():
        with open(css_file, 'r', encoding='utf-8') as f:
            custom_css = f.read()
    
    # Render HTML
    html = template.render(
        title=metadata.get('title', markdown_file.stem),
        metadata=metadata if metadata else None,
        content=html_content,
        custom_css=custom_css,
        generated_date=datetime.now().strftime('%Y-%m-%d %H:%M')
    )
    
    return html


def html_to_docx(markdown_file: Path, output_file: Optional[Path] = None) -> Path:
    """Convert markdown to DOCX using pypandoc"""
    try:
        import pypandoc
    except ImportError:
        click.echo("⚠️  pypandoc를 설치하는 중입니다. 잠시만 기다려주세요...")
        import subprocess
        subprocess.run([sys.executable, "-m", "pip", "install", "pypandoc"], check=True)
        import pypandoc
    
    if output_file is None:
        output_file = markdown_file.with_suffix('.docx')
    
    # Convert directly using pypandoc
    pypandoc.convert_file(
        str(markdown_file),
        'docx',
        outputfile=str(output_file),
        extra_args=['--standalone', '--toc']
    )
    
    return output_file


def extract_tables_to_excel(markdown_file: Path, output_file: Optional[Path] = None) -> Path:
    """Extract tables from markdown and save to Excel"""
    if output_file is None:
        output_file = markdown_file.with_suffix('.xlsx')
    
    # Read and convert markdown
    with open(markdown_file, 'r', encoding='utf-8') as f:
        markdown_text = f.read()
    
    _, content = extract_metadata(markdown_text)
    html_content = markdown2.markdown(content, extras=['tables'])
    
    # Parse HTML to find tables
    soup = BeautifulSoup(html_content, 'html.parser')
    tables = soup.find_all('table')
    
    if not tables:
        click.echo("⚠️  문서에서 표를 찾을 수 없습니다.")
        return None
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    
    for idx, table in enumerate(tables, 1):
        # Create sheet for each table
        sheet_name = f"Table_{idx}"
        ws = wb.create_sheet(title=sheet_name)
        
        # Extract rows
        rows = table.find_all('tr')
        for row_idx, row in enumerate(rows, 1):
            cells = row.find_all(['th', 'td'])
            for col_idx, cell in enumerate(cells, 1):
                cell_value = cell.get_text(strip=True)
                ws.cell(row=row_idx, column=col_idx, value=cell_value)
                
                # Style header row
                if row.find('th'):
                    ws.cell(row=row_idx, column=col_idx).font = header_font
                    ws.cell(row=row_idx, column=col_idx).fill = header_fill
                    ws.cell(row=row_idx, column=col_idx).alignment = header_align
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook
    wb.save(output_file)
    return output_file


@click.command()
@click.argument('markdown_file', type=click.Path(exists=True, path_type=Path))
@click.option('--preview', is_flag=True, help='브라우저에서 미리보기')
@click.option('--docx', is_flag=True, help='DOCX (Word) 파일로 변환')
@click.option('--excel', is_flag=True, help='표를 Excel 파일로 추출')
@click.option('--template', default=None, help='사용할 HTML 템플릿 (report, legal, presentation)')
@click.option('--output', '-o', type=click.Path(), help='출력 파일 경로')
def convert(markdown_file: Path, preview: bool, docx: bool, excel: bool, 
            template: Optional[str], output: Optional[str]):
    """
    연구 문서를 다양한 형식으로 변환합니다.
    
    \b
    사용 예시:
        convert.py report.md --preview     # 브라우저에서 미리보기
        convert.py report.md --docx        # Word 문서로 변환
        convert.py report.md --excel       # 표를 Excel로 추출
    """
    
    # Default to preview if no option specified
    if not any([preview, docx, excel]):
        preview = True
    
    output_path = Path(output) if output else None
    
    try:
        if preview:
            # Generate HTML and open in browser
            html = markdown_to_html(markdown_file, template)
            
            # Save to temporary file
            with tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False, encoding='utf-8') as f:
                f.write(html)
                temp_path = f.name
            
            # Open in browser
            webbrowser.open(f'file://{temp_path}')
            click.echo(f"✅ 브라우저에서 미리보기를 열었습니다: {temp_path}")
            click.echo("💡 PDF로 저장하려면 브라우저에서 Ctrl+P (또는 Cmd+P)를 누르세요")
            
        if docx:
            output_file = html_to_docx(markdown_file, output_path)
            click.echo(f"✅ Word 문서로 변환되었습니다: {output_file}")
            
        if excel:
            output_file = extract_tables_to_excel(markdown_file, output_path)
            if output_file:
                click.echo(f"✅ Excel 파일로 표가 추출되었습니다: {output_file}")
    
    except Exception as e:
        click.echo(f"❌ 오류가 발생했습니다: {str(e)}", err=True)
        sys.exit(1)


if __name__ == '__main__':
    convert()